import openpyxl

wb = openpyxl.load_workbook('Master_Enrichment_Database_Singapore.xlsx')
ws = wb.active

print(f'Sheet name: {ws.title}')
print(f'Dimensions: {ws.dimensions}')
print(f'Total rows: {ws.max_row}')
print(f'Total columns: {ws.max_column}')

print('\nColumn headers:')
headers = [cell.value for cell in ws[1]]
for i, header in enumerate(headers, 1):
    print(f'{i}. {header}')

print('\nFirst 5 data rows:')
for row_num in range(2, min(7, ws.max_row + 1)):
    row_data = [cell.value for cell in ws[row_num]]
    print(f'\nRow {row_num}:')
    for i, (header, value) in enumerate(zip(headers, row_data)):
        if value is not None and value != '':
            print(f'  {header}: {value}')
